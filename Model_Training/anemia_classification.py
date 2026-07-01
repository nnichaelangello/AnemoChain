import os
import re
import sys
import json
import time
import math
import random
import platform
import datetime
import warnings
from pathlib import Path
from collections import Counter

import numpy as np
import pandas as pd
import openpyxl
from PIL import Image
from scipy import stats

from sklearn.svm import SVC
from sklearn.ensemble import RandomForestClassifier, HistGradientBoostingClassifier
from sklearn.neighbors import KNeighborsClassifier
import xgboost as xgb
from sklearn.preprocessing import StandardScaler
from sklearn.pipeline import Pipeline
from sklearn.model_selection import train_test_split
from sklearn.metrics import (
    accuracy_score,
    balanced_accuracy_score,
    precision_score,
    recall_score,
    f1_score,
    roc_auc_score,
    log_loss,
    matthews_corrcoef,
    cohen_kappa_score,
    jaccard_score,
    hamming_loss,
    zero_one_loss,
    brier_score_loss,
    confusion_matrix,
    roc_curve,
    auc,
    precision_recall_curve,
    average_precision_score,
)

warnings.filterwarnings("ignore")

RANDOM_SEED = 42
random.seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)

DATASET_BASE_PATH = Path(r"c:\Users\Michael Angello\Documents\Michael\Lomba\KTI PRISMA\dataset anemia")
OUTPUT_JSON_PATH = Path(r"c:\Users\Michael Angello\Documents\Michael\Lomba\KTI PRISMA\experiment_results.json")

ANEMIA_HGB_THRESHOLD = 12.0
IMAGE_RESIZE_FOR_EXTRACTION = 256
FEATURE_EXTRACTION_MIN_VALID_PIXELS = 500

SPLIT_RATIOS = {
    "90_10": {"train": 0.90, "test": 0.10},
    "80_20": {"train": 0.80, "test": 0.20},
    "70_30": {"train": 0.70, "test": 0.30},
}

EXPERIMENT_TIMESTAMP = datetime.datetime.now().isoformat()


def load_italy_metadata():
    xlsx_path = DATASET_BASE_PATH / "Italy" / "Italy.xlsx"
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        number = row[0]
        hgb_raw = row[1]
        gender = row[2]
        age = row[3]
        if number is None:
            continue
        if hgb_raw is None:
            continue
        if isinstance(hgb_raw, str):
            hgb_cleaned = hgb_raw.replace(",", ".").strip()
            if not re.match(r"^\d+(\.\d+)?$", hgb_cleaned):
                continue
            hgb = float(hgb_cleaned)
        else:
            hgb = float(hgb_raw)
        folder_path = DATASET_BASE_PATH / "Italy" / str(int(number))
        palpebral_files = list(folder_path.glob("*palpebral.png")) if folder_path.exists() else []
        palpebral_files = [f for f in palpebral_files if "forniceal" not in f.name]
        if not palpebral_files:
            continue
        records.append({
            "number": int(number),
            "hgb": hgb,
            "gender": gender,
            "age": float(age) if age is not None else None,
            "source": "Italy",
            "image_path": str(palpebral_files[0]),
        })
    return records


def load_india_metadata():
    xlsx_path = DATASET_BASE_PATH / "India" / "India.xlsx"
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        number = row[0]
        hgb_raw = row[1]
        gender = row[2]
        age = row[3]
        if number is None:
            continue
        if hgb_raw is None:
            continue
        if isinstance(hgb_raw, str):
            hgb_cleaned = hgb_raw.replace(",", ".").strip()
            if not re.match(r"^\d+(\.\d+)?$", hgb_cleaned):
                continue
            hgb = float(hgb_cleaned)
        else:
            hgb = float(hgb_raw)
        folder_path = DATASET_BASE_PATH / "India" / str(int(number))
        palpebral_files = list(folder_path.glob("*palpebral.png")) if folder_path.exists() else []
        palpebral_files = [f for f in palpebral_files if "forniceal" not in f.name]
        if not palpebral_files:
            continue
        records.append({
            "number": int(number),
            "hgb": hgb,
            "gender": gender,
            "age": float(age) if age is not None else None,
            "source": "India",
            "image_path": str(palpebral_files[0]),
        })
    return records


italy_records = load_italy_metadata()
india_records = load_india_metadata()
all_records = italy_records + india_records

print(f"Italy records loaded: {len(italy_records)}")
print(f"India records loaded: {len(india_records)}")
print(f"Total records: {len(all_records)}")


raw_df = pd.DataFrame(all_records)
raw_df["anemia_label"] = (raw_df["hgb"] < ANEMIA_HGB_THRESHOLD).astype(int)
raw_df["anemia_label_name"] = raw_df["anemia_label"].map({0: "Non-Anemia", 1: "Anemia"})

print(raw_df["anemia_label_name"].value_counts())


def compute_target_entropy(class_counts):
    total = sum(class_counts.values())
    entropy = 0.0
    for count in class_counts.values():
        if count > 0:
            proportion = count / total
            entropy -= proportion * math.log2(proportion)
    return entropy


def compute_imbalance_ratio(class_counts):
    counts = list(class_counts.values())
    return max(counts) / min(counts) if min(counts) > 0 else float("inf")


label_counts_raw = raw_df["anemia_label_name"].value_counts().to_dict()
label_proportions_raw = raw_df["anemia_label_name"].value_counts(normalize=True).to_dict()
target_entropy = compute_target_entropy(label_counts_raw)
imbalance_ratio = compute_imbalance_ratio(label_counts_raw)
source_counts_raw = raw_df["source"].value_counts().to_dict()

eda_before_preprocessing = {
    "shape": {"rows": int(raw_df.shape[0]), "columns": int(raw_df.shape[1])},
    "num_rows": int(raw_df.shape[0]),
    "num_columns": int(raw_df.shape[1]),
    "column_names": raw_df.columns.tolist(),
    "dtypes": {col: str(dtype) for col, dtype in raw_df.dtypes.items()},
    "missing_value_count": raw_df.isnull().sum().to_dict(),
    "missing_value_percentage": (raw_df.isnull().sum() / len(raw_df) * 100).to_dict(),
    "duplicate_count": int(raw_df.duplicated().sum()),
    "duplicate_percentage": float(raw_df.duplicated().sum() / len(raw_df) * 100),
    "head_5_rows": raw_df.head(5).to_dict(orient="records"),
    "tail_5_rows": raw_df.tail(5).to_dict(orient="records"),
    "sample_10_rows": raw_df.sample(n=min(10, len(raw_df)), random_state=RANDOM_SEED).to_dict(orient="records"),
    "describe_numeric": raw_df.describe().to_dict(),
    "describe_categorical": raw_df.select_dtypes(include=["object"]).describe().to_dict(),
    "categorical_cardinality": {
        col: int(raw_df[col].nunique())
        for col in raw_df.select_dtypes(include=["object"]).columns
    },
    "target_distribution": {
        "class_counts": {str(k): int(v) for k, v in label_counts_raw.items()},
        "class_proportions": {str(k): float(v) for k, v in label_proportions_raw.items()},
        "num_classes": int(raw_df["anemia_label_name"].nunique()),
        "class_names": raw_df["anemia_label_name"].unique().tolist(),
    },
    "target_entropy": float(target_entropy),
    "imbalance_ratio": float(imbalance_ratio),
    "dataset_info": {
        "total_images": int(len(raw_df)),
        "source_distribution": {str(k): int(v) for k, v in source_counts_raw.items()},
        "hgb_min": float(raw_df["hgb"].min()),
        "hgb_max": float(raw_df["hgb"].max()),
        "hgb_mean": float(raw_df["hgb"].mean()),
        "hgb_std": float(raw_df["hgb"].std()),
        "hgb_median": float(raw_df["hgb"].median()),
        "hgb_q25": float(raw_df["hgb"].quantile(0.25)),
        "hgb_q75": float(raw_df["hgb"].quantile(0.75)),
        "anemia_threshold_gdl": ANEMIA_HGB_THRESHOLD,
        "classification_type": "binary",
        "image_type": "palpebral_conjunctiva_segmented",
        "image_format": "PNG",
        "approach": "multi_colorspace_feature_extraction_plus_ml",
    },
}

print("EDA before preprocessing completed.")


gender_mapping = {"M": 0, "F": 1}
raw_df["gender_encoded"] = raw_df["gender"].map(gender_mapping)

processed_df = raw_df.copy()
processed_df = processed_df.dropna(subset=["hgb", "image_path"])
processed_df = processed_df.reset_index(drop=True)

print(f"After basic preprocessing: {len(processed_df)} samples")


def linearize_srgb_channel(channel_float):
    return np.where(
        channel_float <= 0.04045,
        channel_float / 12.92,
        ((channel_float + 0.055) / 1.055) ** 2.4,
    )


def convert_rgb_to_lab(img_rgb_float64):
    r_lin = linearize_srgb_channel(img_rgb_float64[:, :, 0])
    g_lin = linearize_srgb_channel(img_rgb_float64[:, :, 1])
    b_lin = linearize_srgb_channel(img_rgb_float64[:, :, 2])

    X = 0.4124564 * r_lin + 0.3575761 * g_lin + 0.1804375 * b_lin
    Y = 0.2126729 * r_lin + 0.7151522 * g_lin + 0.0721750 * b_lin
    Z = 0.0193339 * r_lin + 0.1191920 * g_lin + 0.9503041 * b_lin

    xn, yn, zn = 0.95047, 1.00000, 1.08883

    def lab_f(t):
        return np.where(t > 0.008856, t ** (1.0 / 3.0), 7.787 * t + 16.0 / 116.0)

    fx = lab_f(X / xn)
    fy = lab_f(Y / yn)
    fz = lab_f(Z / zn)

    L_star = 116.0 * fy - 16.0
    a_star = 500.0 * (fx - fy)
    b_star = 200.0 * (fy - fz)

    return np.stack([L_star, a_star, b_star], axis=-1)


def convert_rgb_to_hsv(img_rgb_float64):
    R = img_rgb_float64[:, :, 0]
    G = img_rgb_float64[:, :, 1]
    B = img_rgb_float64[:, :, 2]

    maxc = np.maximum(np.maximum(R, G), B)
    minc = np.minimum(np.minimum(R, G), B)
    delta = maxc - minc

    V = maxc
    S = np.where(maxc > 1e-8, delta / maxc, 0.0)

    H = np.zeros_like(R)
    mask_nonzero_delta = delta > 1e-8
    mask_r = mask_nonzero_delta & (maxc == R)
    mask_g = mask_nonzero_delta & (maxc == G)
    mask_b = mask_nonzero_delta & (maxc == B)

    H[mask_r] = ((G[mask_r] - B[mask_r]) / delta[mask_r]) % 6.0
    H[mask_g] = (B[mask_g] - R[mask_g]) / delta[mask_g] + 2.0
    H[mask_b] = (R[mask_b] - G[mask_b]) / delta[mask_b] + 4.0

    H = (H / 6.0) % 1.0

    return np.stack([H, S, V], axis=-1)


def convert_rgb_to_ycbcr(img_rgb_float64):
    R = img_rgb_float64[:, :, 0]
    G = img_rgb_float64[:, :, 1]
    B = img_rgb_float64[:, :, 2]

    Y  =  0.299000 * R + 0.587000 * G + 0.114000 * B
    Cb = -0.168736 * R - 0.331264 * G + 0.500000 * B + 0.5
    Cr =  0.500000 * R - 0.418688 * G - 0.081312 * B + 0.5

    return np.stack([Y, Cb, Cr], axis=-1)


def compute_channel_statistics(pixel_values_1d, prefix):
    result = {}
    result[f"{prefix}_mean"] = float(np.mean(pixel_values_1d))
    result[f"{prefix}_std"] = float(np.std(pixel_values_1d))
    result[f"{prefix}_skewness"] = float(stats.skew(pixel_values_1d))
    result[f"{prefix}_kurtosis"] = float(stats.kurtosis(pixel_values_1d))
    result[f"{prefix}_p10"] = float(np.percentile(pixel_values_1d, 10))
    result[f"{prefix}_p25"] = float(np.percentile(pixel_values_1d, 25))
    result[f"{prefix}_p50"] = float(np.percentile(pixel_values_1d, 50))
    result[f"{prefix}_p75"] = float(np.percentile(pixel_values_1d, 75))
    result[f"{prefix}_p90"] = float(np.percentile(pixel_values_1d, 90))
    return result


def extract_conjunctiva_features(image_path):
    img_pil = Image.open(image_path)

    if img_pil.mode == "RGBA":
        img_resized = img_pil.resize(
            (IMAGE_RESIZE_FOR_EXTRACTION, IMAGE_RESIZE_FOR_EXTRACTION), Image.LANCZOS
        )
        img_array = np.array(img_resized, dtype=np.float64)
        valid_mask = img_array[:, :, 3] > 10.0
        img_rgb_float = img_array[:, :, :3] / 255.0
    else:
        img_rgb_pil = img_pil.convert("RGB").resize(
            (IMAGE_RESIZE_FOR_EXTRACTION, IMAGE_RESIZE_FOR_EXTRACTION), Image.LANCZOS
        )
        img_rgb_float = np.array(img_rgb_pil, dtype=np.float64) / 255.0
        luminance = (
            0.299 * img_rgb_float[:, :, 0]
            + 0.587 * img_rgb_float[:, :, 1]
            + 0.114 * img_rgb_float[:, :, 2]
        )
        valid_mask = luminance > 0.04

    num_valid_pixels = int(valid_mask.sum())

    if num_valid_pixels < FEATURE_EXTRACTION_MIN_VALID_PIXELS:
        raise ValueError(f"Insufficient valid pixels: {num_valid_pixels}")

    img_lab = convert_rgb_to_lab(img_rgb_float)
    img_hsv = convert_rgb_to_hsv(img_rgb_float)
    img_ycbcr = convert_rgb_to_ycbcr(img_rgb_float)

    R = img_rgb_float[:, :, 0][valid_mask]
    G = img_rgb_float[:, :, 1][valid_mask]
    B = img_rgb_float[:, :, 2][valid_mask]

    L_star = img_lab[:, :, 0][valid_mask]
    a_star = img_lab[:, :, 1][valid_mask]
    b_star = img_lab[:, :, 2][valid_mask]

    H_channel = img_hsv[:, :, 0][valid_mask]
    S_channel = img_hsv[:, :, 1][valid_mask]
    V_channel = img_hsv[:, :, 2][valid_mask]

    Y_channel = img_ycbcr[:, :, 0][valid_mask]
    Cb_channel = img_ycbcr[:, :, 1][valid_mask]
    Cr_channel = img_ycbcr[:, :, 2][valid_mask]

    features = {}

    for channel_data, channel_name in [
        (R, "rgb_R"),
        (G, "rgb_G"),
        (B, "rgb_B"),
    ]:
        features.update(compute_channel_statistics(channel_data, channel_name))

    for channel_data, channel_name in [
        (L_star, "lab_L"),
        (a_star, "lab_a"),
        (b_star, "lab_b"),
    ]:
        features.update(compute_channel_statistics(channel_data, channel_name))

    for channel_data, channel_name in [
        (H_channel, "hsv_H"),
        (S_channel, "hsv_S"),
        (V_channel, "hsv_V"),
    ]:
        features.update(compute_channel_statistics(channel_data, channel_name))

    for channel_data, channel_name in [
        (Y_channel, "ycbcr_Y"),
        (Cb_channel, "ycbcr_Cb"),
        (Cr_channel, "ycbcr_Cr"),
    ]:
        features.update(compute_channel_statistics(channel_data, channel_name))

    epsilon = 1e-8

    redness_index = R / (G + B + epsilon)
    features.update(compute_channel_statistics(redness_index, "pallor_redness_index"))

    norm_total = R + G + B + epsilon
    norm_red = R / norm_total
    norm_green = G / norm_total
    norm_blue = B / norm_total
    features.update(compute_channel_statistics(norm_red, "pallor_norm_red"))
    features.update(compute_channel_statistics(norm_green, "pallor_norm_green"))
    features.update(compute_channel_statistics(norm_blue, "pallor_norm_blue"))

    chroma_lab = np.sqrt(a_star ** 2 + b_star ** 2)
    features.update(compute_channel_statistics(chroma_lab, "pallor_chroma_lab"))

    hue_angle_lab_degrees = np.degrees(np.arctan2(b_star, a_star))
    features.update(compute_channel_statistics(hue_angle_lab_degrees, "pallor_hue_angle_lab"))

    green_minus_red = G - R
    features.update(compute_channel_statistics(green_minus_red, "pallor_green_minus_red"))

    features["meta_num_valid_pixels"] = num_valid_pixels
    features["meta_valid_pixel_ratio"] = float(valid_mask.sum() / valid_mask.size)

    return features


print("Extracting features from all conjunctiva images...")
extraction_start_time = time.time()

all_feature_dicts = []
extraction_failed_indices = []
extraction_errors = {}

for sample_idx in range(len(processed_df)):
    current_image_path = processed_df.iloc[sample_idx]["image_path"]
    try:
        feature_dict = extract_conjunctiva_features(current_image_path)
        all_feature_dicts.append(feature_dict)
    except Exception as extraction_error:
        extraction_failed_indices.append(sample_idx)
        extraction_errors[sample_idx] = str(extraction_error)

total_extraction_time = time.time() - extraction_start_time

valid_sample_indices = [i for i in range(len(processed_df)) if i not in extraction_failed_indices]
processed_df = processed_df.iloc[valid_sample_indices].reset_index(drop=True)

feature_df_full = pd.DataFrame(all_feature_dicts)

META_COLUMNS = [col for col in feature_df_full.columns if col.startswith("meta_")]
FEATURE_COLUMNS = [col for col in feature_df_full.columns if not col.startswith("meta_")]

feature_df = feature_df_full[FEATURE_COLUMNS].copy()
meta_df = feature_df_full[META_COLUMNS].copy()

X = feature_df.values
all_labels = processed_df["anemia_label"].tolist()
all_image_paths = processed_df["image_path"].tolist()
all_indices = list(range(len(processed_df)))

CLASS_NAMES = ["Non-Anemia", "Anemia"]
NUM_CLASSES = 2

print(f"Feature extraction complete. {len(processed_df)} samples, {len(FEATURE_COLUMNS)} features.")
print(f"Total extraction time: {total_extraction_time:.2f}s")
print(f"Failed extractions: {len(extraction_failed_indices)}")


final_label_counts = processed_df["anemia_label_name"].value_counts().to_dict()
final_label_proportions = processed_df["anemia_label_name"].value_counts(normalize=True).to_dict()

y_array_for_corr = np.array(all_labels)
feature_target_correlations = {}
for feat_name in FEATURE_COLUMNS:
    feat_values = feature_df[feat_name].values
    corr_result = stats.pointbiserialr(y_array_for_corr, feat_values)
    feature_target_correlations[feat_name] = {
        "point_biserial_r": float(corr_result.statistic),
        "p_value": float(corr_result.pvalue),
        "abs_r": float(abs(corr_result.statistic)),
    }

top_features_by_correlation = sorted(
    feature_target_correlations.items(),
    key=lambda x: x[1]["abs_r"],
    reverse=True,
)[:30]

eda_after_preprocessing = {
    "shape_after_preprocessing": {"rows": int(processed_df.shape[0]), "columns": len(FEATURE_COLUMNS)},
    "num_features_final": len(FEATURE_COLUMNS),
    "feature_names_final": FEATURE_COLUMNS,
    "feature_groups": {
        "rgb_statistics": [f for f in FEATURE_COLUMNS if f.startswith("rgb_")],
        "lab_statistics": [f for f in FEATURE_COLUMNS if f.startswith("lab_")],
        "hsv_statistics": [f for f in FEATURE_COLUMNS if f.startswith("hsv_")],
        "ycbcr_statistics": [f for f in FEATURE_COLUMNS if f.startswith("ycbcr_")],
        "pallor_indices": [f for f in FEATURE_COLUMNS if f.startswith("pallor_")],
    },
    "feature_descriptive_statistics": feature_df.describe().to_dict(),
    "missing_values_after_preprocessing": feature_df.isnull().sum().to_dict(),
    "target_distribution_after_preprocessing": {
        "class_counts": {str(k): int(v) for k, v in final_label_counts.items()},
        "class_proportions": {str(k): float(v) for k, v in final_label_proportions.items()},
    },
    "feature_target_correlations": {
        k: v for k, v in feature_target_correlations.items()
    },
    "top_30_features_by_abs_correlation": [
        {"feature": feat, **corr_info}
        for feat, corr_info in top_features_by_correlation
    ],
    "image_quality_metadata": {
        "mean_valid_pixels": float(meta_df["meta_num_valid_pixels"].mean()),
        "std_valid_pixels": float(meta_df["meta_num_valid_pixels"].std()),
        "min_valid_pixels": float(meta_df["meta_num_valid_pixels"].min()),
        "max_valid_pixels": float(meta_df["meta_num_valid_pixels"].max()),
        "mean_valid_pixel_ratio": float(meta_df["meta_valid_pixel_ratio"].mean()),
    },
    "transformation_info": {
        "image_resize_for_extraction": IMAGE_RESIZE_FOR_EXTRACTION,
        "masking_strategy": "alpha_channel_for_RGBA_luminance_threshold_for_RGB",
        "luminance_threshold": 0.04,
        "alpha_threshold": 10,
        "statistics_computed_per_channel": [
            "mean", "std", "skewness", "kurtosis",
            "p10", "p25", "p50", "p75", "p90",
        ],
        "color_spaces_used": ["RGB", "CIE_Lab", "HSV", "YCbCr"],
        "derived_indices": [
            "redness_index (R/(G+B))",
            "normalized_red (R/(R+G+B))",
            "normalized_green (G/(R+G+B))",
            "normalized_blue (B/(R+G+B))",
            "chroma_lab (sqrt(a*^2+b*^2))",
            "hue_angle_lab (atan2(b*,a*) degrees)",
            "green_minus_red (G-R pallor indicator)",
        ],
    },
    "extraction_info": {
        "total_extraction_time_seconds": float(total_extraction_time),
        "extraction_time_per_sample_seconds": float(total_extraction_time / len(all_feature_dicts)) if all_feature_dicts else 0,
        "failed_extractions": len(extraction_failed_indices),
        "failed_extraction_errors": extraction_errors,
    },
    "encoding_mapping": {
        "gender": gender_mapping,
        "anemia_label": {"Non-Anemia": 0, "Anemia": 1},
    },
    "scaler_info": {
        "type": "StandardScaler fitted only on training fold (no data leakage)",
        "note": "Each model uses a Pipeline: StandardScaler -> Classifier. Scaler is fit on train set only.",
    },
    "records_removed_basic_preprocessing": int(len(raw_df) - len(valid_sample_indices) - len(extraction_failed_indices)),
    "records_removed_extraction_failure": len(extraction_failed_indices),
    "final_dataset_size": int(len(processed_df)),
}

print(f"EDA after preprocessing completed. Top correlated feature: {top_features_by_correlation[0][0]}")


split_data_info = {}

for split_name, split_ratio in SPLIT_RATIOS.items():
    train_indices, test_indices = train_test_split(
        all_indices,
        test_size=split_ratio["test"],
        random_state=RANDOM_SEED,
        stratify=all_labels,
    )
    train_labels_split = [all_labels[i] for i in train_indices]
    test_labels_split = [all_labels[i] for i in test_indices]
    train_label_counts = Counter(train_labels_split)
    test_label_counts = Counter(test_labels_split)
    train_total = len(train_indices)
    test_total = len(test_indices)
    split_data_info[split_name] = {
        "num_train": train_total,
        "num_test": test_total,
        "train_indices": train_indices,
        "test_indices": test_indices,
        "train_class_distribution": {str(k): int(v) for k, v in train_label_counts.items()},
        "test_class_distribution": {str(k): int(v) for k, v in test_label_counts.items()},
        "train_class_proportions": {str(k): float(v / train_total) for k, v in train_label_counts.items()},
        "test_class_proportions": {str(k): float(v / test_total) for k, v in test_label_counts.items()},
        "train_ratio": split_ratio["train"],
        "test_ratio": split_ratio["test"],
    }

print("Data split information computed.")


def build_svm_pipeline():
    svm_classifier = SVC(
        kernel="rbf",
        C=1.0,
        gamma="scale",
        probability=True,
        class_weight="balanced",
        random_state=RANDOM_SEED,
    )
    return Pipeline([("scaler", StandardScaler()), ("classifier", svm_classifier)])


def build_random_forest_pipeline():
    rf_classifier = RandomForestClassifier(
        n_estimators=500,
        max_depth=None,
        min_samples_split=5,
        min_samples_leaf=2,
        max_features="sqrt",
        class_weight="balanced",
        random_state=RANDOM_SEED,
        n_jobs=-1,
    )
    return Pipeline([("scaler", StandardScaler()), ("classifier", rf_classifier)])


def build_hist_gradient_boosting_pipeline():
    hgbc_classifier = HistGradientBoostingClassifier(
        max_iter=300,
        max_leaf_nodes=31,
        min_samples_leaf=10,
        learning_rate=0.05,
        class_weight="balanced",
        random_state=RANDOM_SEED,
    )
    return Pipeline([("scaler", StandardScaler()), ("classifier", hgbc_classifier)])


def build_knn_pipeline():
    knn_classifier = KNeighborsClassifier(
        n_neighbors=5,
        weights="distance",
        n_jobs=-1,
    )
    return Pipeline([("scaler", StandardScaler()), ("classifier", knn_classifier)])


def build_xgboost_pipeline():
    xgb_classifier = xgb.XGBClassifier(
        n_estimators=300,
        max_depth=6,
        learning_rate=0.05,
        random_state=RANDOM_SEED,
        eval_metric="logloss",
        n_jobs=-1,
    )
    return Pipeline([("scaler", StandardScaler()), ("classifier", xgb_classifier)])


model_builders = {
    "SVM_RBF": build_svm_pipeline,
    "RandomForest": build_random_forest_pipeline,
    "HistGradientBoosting": build_hist_gradient_boosting_pipeline,
    "KNN": build_knn_pipeline,
    "XGBoost": build_xgboost_pipeline,
}

print(f"Models defined: {list(model_builders.keys())}")


def compute_binary_metrics(y_true, y_pred, y_proba_list, class_names):
    y_true_np = np.array(y_true)
    y_pred_np = np.array(y_pred)
    y_proba_np = np.array(y_proba_list)
    y_proba_positive = y_proba_np[:, 1]

    cm = confusion_matrix(y_true_np, y_pred_np)
    tn, fp, fn, tp = cm.ravel()

    sensitivity = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    specificity = tn / (tn + fp) if (tn + fp) > 0 else 0.0
    tpr = sensitivity
    fpr_val = fp / (fp + tn) if (fp + tn) > 0 else 0.0
    tnr = specificity
    fnr = fn / (fn + tp) if (fn + tp) > 0 else 0.0

    roc_fpr, roc_tpr, roc_thresholds = roc_curve(y_true_np, y_proba_positive)
    roc_auc_value = auc(roc_fpr, roc_tpr)

    pr_precision_vals, pr_recall_vals, pr_thresholds = precision_recall_curve(y_true_np, y_proba_positive)
    avg_precision_value = average_precision_score(y_true_np, y_proba_positive)

    metrics = {
        "accuracy": float(accuracy_score(y_true_np, y_pred_np)),
        "balanced_accuracy": float(balanced_accuracy_score(y_true_np, y_pred_np)),
        "precision_macro": float(precision_score(y_true_np, y_pred_np, average="macro", zero_division=0)),
        "precision_micro": float(precision_score(y_true_np, y_pred_np, average="micro", zero_division=0)),
        "precision_weighted": float(precision_score(y_true_np, y_pred_np, average="weighted", zero_division=0)),
        "recall_macro": float(recall_score(y_true_np, y_pred_np, average="macro", zero_division=0)),
        "recall_micro": float(recall_score(y_true_np, y_pred_np, average="micro", zero_division=0)),
        "recall_weighted": float(recall_score(y_true_np, y_pred_np, average="weighted", zero_division=0)),
        "f1_macro": float(f1_score(y_true_np, y_pred_np, average="macro", zero_division=0)),
        "f1_micro": float(f1_score(y_true_np, y_pred_np, average="micro", zero_division=0)),
        "f1_weighted": float(f1_score(y_true_np, y_pred_np, average="weighted", zero_division=0)),
        "roc_auc": float(roc_auc_score(y_true_np, y_proba_positive)),
        "roc_auc_ovr": float(roc_auc_score(y_true_np, y_proba_positive)),
        "roc_auc_ovo": float(roc_auc_score(y_true_np, y_proba_positive)),
        "log_loss": float(log_loss(y_true_np, y_proba_np)),
        "matthews_corrcoef": float(matthews_corrcoef(y_true_np, y_pred_np)),
        "cohen_kappa": float(cohen_kappa_score(y_true_np, y_pred_np)),
        "jaccard_score_macro": float(jaccard_score(y_true_np, y_pred_np, average="macro", zero_division=0)),
        "jaccard_score_weighted": float(jaccard_score(y_true_np, y_pred_np, average="weighted", zero_division=0)),
        "hamming_loss": float(hamming_loss(y_true_np, y_pred_np)),
        "zero_one_loss": float(zero_one_loss(y_true_np, y_pred_np)),
        "brier_score": float(brier_score_loss(y_true_np, y_proba_positive)),
        "sensitivity": float(sensitivity),
        "specificity": float(specificity),
        "true_positive_rate": float(tpr),
        "false_positive_rate": float(fpr_val),
        "true_negative_rate": float(tnr),
        "false_negative_rate": float(fnr),
        "per_class": {
            class_names[cls_idx]: {
                "precision": float(
                    precision_score(y_true_np, y_pred_np, labels=[cls_idx], average=None, zero_division=0)[0]
                ),
                "recall": float(
                    recall_score(y_true_np, y_pred_np, labels=[cls_idx], average=None, zero_division=0)[0]
                ),
                "f1_score": float(
                    f1_score(y_true_np, y_pred_np, labels=[cls_idx], average=None, zero_division=0)[0]
                ),
            }
            for cls_idx in range(len(class_names))
        },
        "confusion_matrix": {
            "raw": cm.tolist(),
            "tn": int(tn),
            "fp": int(fp),
            "fn": int(fn),
            "tp": int(tp),
        },
        "roc_analysis": {
            "fpr": roc_fpr.tolist(),
            "tpr": roc_tpr.tolist(),
            "thresholds": roc_thresholds.tolist(),
            "auc": float(roc_auc_value),
        },
        "precision_recall_analysis": {
            "precision": pr_precision_vals.tolist(),
            "recall": pr_recall_vals.tolist(),
            "thresholds": pr_thresholds.tolist(),
            "average_precision": float(avg_precision_value),
        },
    }
    return metrics


def train_and_evaluate_ml_model(
    model_name, builder_fn, split_name,
    X_train, X_test, y_train, y_test, feature_names
):
    pipeline = builder_fn()

    train_start = time.time()
    pipeline.fit(X_train, np.array(y_train))
    training_time = time.time() - train_start

    train_inf_start = time.time()
    train_y_proba = pipeline.predict_proba(X_train)
    train_y_pred = pipeline.predict(X_train)
    train_inference_time = time.time() - train_inf_start

    test_inf_start = time.time()
    test_y_proba = pipeline.predict_proba(X_test)
    test_y_pred = pipeline.predict(X_test)
    test_inference_time = time.time() - test_inf_start

    fitted_classifier = pipeline.named_steps["classifier"]
    fitted_scaler = pipeline.named_steps["scaler"]

    training_info = {
        "model_name": model_name,
        "model_params": {str(k): str(v) for k, v in fitted_classifier.get_params().items()},
        "pipeline_steps": list(pipeline.named_steps.keys()),
        "n_features_used": len(feature_names),
        "feature_names": feature_names,
        "training_time_seconds": float(training_time),
        "class_weight_strategy": "balanced",
        "scaler_params": {
            "mean_": fitted_scaler.mean_.tolist(),
            "scale_": fitted_scaler.scale_.tolist(),
            "var_": fitted_scaler.var_.tolist(),
        },
    }

    if hasattr(fitted_classifier, "feature_importances_"):
        importances = fitted_classifier.feature_importances_
        importance_dict = {feat: float(imp) for feat, imp in zip(feature_names, importances)}
        sorted_importances = sorted(importance_dict.items(), key=lambda x: x[1], reverse=True)
        training_info["feature_importances"] = importance_dict
        training_info["feature_importances_top_20"] = dict(sorted_importances[:20])

    if model_name == "SVM_RBF":
        training_info["svm_n_support_per_class"] = fitted_classifier.n_support_.tolist()
        training_info["svm_total_support_vectors"] = int(fitted_classifier.n_support_.sum())
        training_info["feature_importances_note"] = (
            "SVM with RBF kernel does not produce native feature importances. "
            "Support vector count is provided instead."
        )

    train_decision_scores = None
    test_decision_scores = None
    if model_name == "SVM_RBF" and hasattr(pipeline, "decision_function"):
        train_decision_scores = pipeline.decision_function(X_train).tolist()
        test_decision_scores = pipeline.decision_function(X_test).tolist()

    train_metrics = compute_binary_metrics(
        list(y_train), list(train_y_pred), train_y_proba.tolist(), CLASS_NAMES
    )
    test_metrics = compute_binary_metrics(
        list(y_test), list(test_y_pred), test_y_proba.tolist(), CLASS_NAMES
    )

    train_confidence = [max(p) for p in train_y_proba.tolist()]
    test_confidence = [max(p) for p in test_y_proba.tolist()]

    return {
        "model_name": model_name,
        "split_name": split_name,
        "training_info": training_info,
        "inference_info": {
            "train_total_inference_time_seconds": float(train_inference_time),
            "train_inference_time_per_sample_seconds": float(train_inference_time / len(y_train)),
            "test_total_inference_time_seconds": float(test_inference_time),
            "test_inference_time_per_sample_seconds": float(test_inference_time / len(y_test)),
        },
        "train_metrics": train_metrics,
        "test_metrics": test_metrics,
        "train_predictions": {
            "y_true": list(y_train),
            "y_pred": list(train_y_pred),
            "y_proba": train_y_proba.tolist(),
            "predicted_class_names": [CLASS_NAMES[p] for p in list(train_y_pred)],
            "prediction_confidence": train_confidence,
            "confidence_mean": float(np.mean(train_confidence)),
            "confidence_std": float(np.std(train_confidence)),
            "confidence_min": float(np.min(train_confidence)),
            "confidence_max": float(np.max(train_confidence)),
            "decision_scores": train_decision_scores,
        },
        "test_predictions": {
            "y_true": list(y_test),
            "y_pred": list(test_y_pred),
            "y_proba": test_y_proba.tolist(),
            "predicted_class_names": [CLASS_NAMES[p] for p in list(test_y_pred)],
            "prediction_confidence": test_confidence,
            "confidence_mean": float(np.mean(test_confidence)),
            "confidence_std": float(np.std(test_confidence)),
            "confidence_min": float(np.min(test_confidence)),
            "confidence_max": float(np.max(test_confidence)),
            "decision_scores": test_decision_scores,
        },
    }


all_experiment_results = {}

for split_name, split_info in split_data_info.items():
    train_idx = split_info["train_indices"]
    test_idx = split_info["test_indices"]
    X_train = X[train_idx]
    X_test = X[test_idx]
    y_train = [all_labels[i] for i in train_idx]
    y_test = [all_labels[i] for i in test_idx]
    all_experiment_results[split_name] = {}
    for model_name, builder_fn in model_builders.items():
        print(f"\nTraining {model_name} with split {split_name}...")
        result = train_and_evaluate_ml_model(
            model_name, builder_fn, split_name,
            X_train, X_test, y_train, y_test, FEATURE_COLUMNS,
        )
        all_experiment_results[split_name][model_name] = result
        print(f"  Test Accuracy:    {result['test_metrics']['accuracy']:.6f}")
        print(f"  Test F1 Macro:    {result['test_metrics']['f1_macro']:.6f}")
        print(f"  Test ROC AUC:     {result['test_metrics']['roc_auc']:.6f}")
        print(f"  Test Sensitivity: {result['test_metrics']['sensitivity']:.6f}")
        print(f"  Test Specificity: {result['test_metrics']['specificity']:.6f}")


def get_library_version(library_name):
    try:
        module = __import__(library_name)
        return getattr(module, "__version__", "unknown")
    except ImportError:
        return "not_installed"


library_versions = {
    "numpy": np.__version__,
    "pandas": pd.__version__,
    "scipy": get_library_version("scipy"),
    "scikit_learn": get_library_version("sklearn"),
    "pillow": get_library_version("PIL"),
    "openpyxl": get_library_version("openpyxl"),
    "python": sys.version,
}

try:
    import sklearn
    library_versions["scikit_learn"] = sklearn.__version__
except ImportError:
    pass

try:
    import PIL
    library_versions["pillow"] = PIL.__version__
except ImportError:
    pass

hardware_info = {
    "cpu": platform.processor(),
    "cpu_cores_logical": os.cpu_count(),
    "platform": platform.platform(),
    "architecture": platform.machine(),
}

try:
    import psutil
    hardware_info["ram_total_gb"] = round(psutil.virtual_memory().total / (1024 ** 3), 4)
except ImportError:
    hardware_info["ram_total_gb"] = "psutil not installed"

environment_metadata = {
    "python_version": sys.version,
    "library_versions": library_versions,
    "os_name": platform.system(),
    "os_version": platform.version(),
    "os_release": platform.release(),
    "hardware": hardware_info,
    "random_seed": RANDOM_SEED,
    "experiment_timestamp": EXPERIMENT_TIMESTAMP,
    "compute_device": "CPU (scikit-learn ML approach)",
}


final_json = {
    "metadata": {
        "experiment_name": "Anemia Detection from Palpebral Conjunctiva - Color Feature Extraction + ML",
        "experiment_description": (
            "Binary classification of anemia vs non-anemia using multi-colorspace "
            "color feature extraction from segmented palpebral conjunctiva images "
            "(Eyes-defy-anemia dataset). Features extracted from RGB, CIE L*a*b*, "
            "HSV, YCbCr color spaces plus derived clinical pallor indices. "
            f"Anemia defined as Hgb < {ANEMIA_HGB_THRESHOLD} g/dL (WHO standard). "
            "Models: SVM-RBF, RandomForest, HistGradientBoosting."
        ),
        "clinical_rationale": (
            "Anemia causes reduction in hemoglobin concentration, resulting in "
            "pallor (decreased redness and saturation) of the palpebral conjunctiva. "
            "Color features directly quantify this pallor, making them the most "
            "clinically aligned representation for this diagnostic task."
        ),
        "dataset_name": "Eyes-defy-anemia",
        "dataset_sources": ["Italy", "India"],
        "classification_type": "binary",
        "class_names": CLASS_NAMES,
        "num_classes": NUM_CLASSES,
        "anemia_hgb_threshold_gdl": ANEMIA_HGB_THRESHOLD,
        "image_type_used": "palpebral conjunctiva (pre-segmented, palpebral.png)",
        "image_resize_for_extraction": IMAGE_RESIZE_FOR_EXTRACTION,
        "num_features_total": len(FEATURE_COLUMNS),
        "feature_groups_summary": {
            "rgb_features": len([f for f in FEATURE_COLUMNS if f.startswith("rgb_")]),
            "lab_features": len([f for f in FEATURE_COLUMNS if f.startswith("lab_")]),
            "hsv_features": len([f for f in FEATURE_COLUMNS if f.startswith("hsv_")]),
            "ycbcr_features": len([f for f in FEATURE_COLUMNS if f.startswith("ycbcr_")]),
            "pallor_index_features": len([f for f in FEATURE_COLUMNS if f.startswith("pallor_")]),
        },
        "models_used": list(model_builders.keys()),
        "splits_evaluated": list(SPLIT_RATIOS.keys()),
        "environment": environment_metadata,
    },
    "eda_before_preprocessing": eda_before_preprocessing,
    "eda_after_preprocessing": eda_after_preprocessing,
    "split_info": {
        split_name: {
            k: v for k, v in split_info.items()
            if k not in ["train_indices", "test_indices"]
        }
        for split_name, split_info in split_data_info.items()
    },
    "experiment_results": all_experiment_results,
}


with open(OUTPUT_JSON_PATH, "w", encoding="utf-8") as output_file:
    json.dump(final_json, output_file, ensure_ascii=False, indent=2, default=str)

print(f"\nResults saved to: {OUTPUT_JSON_PATH}")
print(f"File size: {OUTPUT_JSON_PATH.stat().st_size / (1024 * 1024):.2f} MB")


print("\n" + "=" * 65)
print("EXPERIMENT SUMMARY")
print("=" * 65)

for split_name in SPLIT_RATIOS.keys():
    print(f"\n[Split: {split_name}]")
    for model_name in model_builders.keys():
        result = all_experiment_results[split_name][model_name]
        test_m = result["test_metrics"]
        print(
            f"  {model_name:<25} | "
            f"Acc: {test_m['accuracy']:.4f} | "
            f"F1:  {test_m['f1_macro']:.4f} | "
            f"AUC: {test_m['roc_auc']:.4f} | "
            f"Sens: {test_m['sensitivity']:.4f} | "
            f"Spec: {test_m['specificity']:.4f}"
        )


print("\n" + "=" * 65)
print("FINAL PRODUCTION MODEL: RETRAIN ON 100% DATA + EXPORT")
print("=" * 65)

FINAL_CLS_JOBLIB_PATH = Path(r"c:\Users\Michael Angello\Documents\Michael\Lomba\KTI PRISMA\final_anemia_classifier.joblib")
FINAL_CLS_METADATA_PATH = Path(r"c:\Users\Michael Angello\Documents\Michael\Lomba\KTI PRISMA\final_anemia_classifier_metadata.json")
FINAL_CLS_ONNX_PATH = Path(r"c:\Users\Michael Angello\Documents\Michael\Lomba\KTI PRISMA\final_anemia_classifier.onnx")

model_avg_roc_auc = {}
for candidate_model_name in model_builders.keys():
    auc_values = [
        all_experiment_results[split_name][candidate_model_name]["test_metrics"]["roc_auc"]
        for split_name in SPLIT_RATIOS.keys()
    ]
    model_avg_roc_auc[candidate_model_name] = float(np.mean(auc_values))

best_final_model_name = max(model_avg_roc_auc, key=lambda k: model_avg_roc_auc[k])
best_final_model_avg_auc = model_avg_roc_auc[best_final_model_name]

print(f"\nModel selection by average test ROC AUC across {len(SPLIT_RATIOS)} splits:")
for mn, avg_auc in sorted(model_avg_roc_auc.items(), key=lambda x: x[1], reverse=True):
    marker = " <-- SELECTED" if mn == best_final_model_name else ""
    print(f"  {mn:<25}: {avg_auc:.4f}{marker}")

print(f"\nRetraining {best_final_model_name} on 100% of data ({len(all_labels)} samples)...")

final_classification_pipeline = model_builders[best_final_model_name]()
final_cls_train_start = time.time()
final_classification_pipeline.fit(X, np.array(all_labels))
final_cls_train_time = time.time() - final_cls_train_start

print(f"Retraining complete in {final_cls_train_time:.2f}s")

import joblib
joblib.dump(final_classification_pipeline, str(FINAL_CLS_JOBLIB_PATH))
print(f"Model saved (joblib): {FINAL_CLS_JOBLIB_PATH}")

final_cls_fitted = final_classification_pipeline.named_steps["classifier"]
final_cls_scaler = final_classification_pipeline.named_steps["scaler"]

final_cls_feature_importances = None
if hasattr(final_cls_fitted, "feature_importances_"):
    importances = final_cls_fitted.feature_importances_
    final_cls_feature_importances = dict(
        sorted(
            {f: float(v) for f, v in zip(FEATURE_COLUMNS, importances)}.items(),
            key=lambda x: x[1],
            reverse=True,
        )
    )

final_cls_metadata = {
    "model_name": best_final_model_name,
    "selection_criterion": "highest average test ROC AUC across all splits",
    "model_avg_roc_auc_all_candidates": {str(k): float(v) for k, v in model_avg_roc_auc.items()},
    "training_data_size": len(all_labels),
    "training_time_seconds": float(final_cls_train_time),
    "class_names": CLASS_NAMES,
    "class_mapping": {"0": "Non-Anemia", "1": "Anemia"},
    "anemia_threshold_gdl": ANEMIA_HGB_THRESHOLD,
    "feature_names_ordered": FEATURE_COLUMNS,
    "n_features": len(FEATURE_COLUMNS),
    "feature_groups": {
        "rgb_statistics": [f for f in FEATURE_COLUMNS if f.startswith("rgb_")],
        "lab_statistics": [f for f in FEATURE_COLUMNS if f.startswith("lab_")],
        "hsv_statistics": [f for f in FEATURE_COLUMNS if f.startswith("hsv_")],
        "ycbcr_statistics": [f for f in FEATURE_COLUMNS if f.startswith("ycbcr_")],
        "pallor_indices": [f for f in FEATURE_COLUMNS if f.startswith("pallor_")],
    },
    "model_params": {str(k): str(v) for k, v in final_cls_fitted.get_params().items()},
    "scaler_params": {
        "mean_": final_cls_scaler.mean_.tolist(),
        "scale_": final_cls_scaler.scale_.tolist(),
        "var_": final_cls_scaler.var_.tolist(),
    },
    "feature_importances_sorted": final_cls_feature_importances,
    "model_files": {
        "joblib": str(FINAL_CLS_JOBLIB_PATH),
        "onnx": str(FINAL_CLS_ONNX_PATH),
    },
    "mobile_inference_pipeline": [
        "1. [UI] Tampilkan panduan capture: overlay frame guide + instruksi tarik kelopak bawah",
        "2. [UI] User mengisi frame dengan konjungvita palpebral secara close-up",
        "3. [Gate 1] MediaPipe eye detection: validasi bahwa ada mata dalam frame, tolak jika tidak terdeteksi",
        "4. [Gate 2] Color distribution check: hitung mean a* (Lab) dan S (HSV) dari piksel. Tolak jika jaringan tidak menunjukkan karakteristik warna konjungvita (bukan pinkish)",
        "5. [Masking] Hitung luminance mask: valid_mask = (0.299*R + 0.587*G + 0.114*B) / 255.0 > 0.04. Gunakan hanya piksel valid untuk ekstraksi fitur",
        "6. [Feature Extraction] Ekstrak 162 fitur warna dari piksel valid (urutan sesuai feature_names_ordered): RGB, CIE L*a*b*, HSV, YCbCr statistics + pallor indices",
        "7. [Preprocessing] Terapkan StandardScaler: scaled[i] = (feature[i] - mean_[i]) / scale_[i]",
        "8. [Inference] Jalankan predict_proba() dari model yang di-load (joblib atau ONNX)",
        "9. [Output] Tampilkan prediksi (Non-Anemia / Anemia), confidence score, dan indikator klinis (redness index, chroma, saturation)",
    ],

}

onnx_export_status = "not_attempted"
try:
    from skl2onnx import convert_sklearn
    from skl2onnx.common.data_types import FloatTensorType
    from skl2onnx import update_registered_converter
    
    # Register XGBoost converter for ONNX
    try:
        from onnxmltools.convert.xgboost.shape_calculators.Classifier import calculate_xgboost_classifier_output_shapes
        from onnxmltools.convert.xgboost.operator_converters.XGBoost import convert_xgboost
        update_registered_converter(
            xgb.XGBClassifier, 'XGBoostXGBClassifier',
            calculate_xgboost_classifier_output_shapes, convert_xgboost,
            options={'nocl': [True, False], 'zipmap': [True, False, 'columns']}
        )
    except Exception as e:
        print(f"Warning: Could not register XGBoost ONNX converter: {e}")

    initial_type = [("float_input", FloatTensorType([None, len(FEATURE_COLUMNS)]))]
    onnx_model_proto = convert_sklearn(
        final_classification_pipeline, initial_types=initial_type, target_opset=12
    )
    with open(FINAL_CLS_ONNX_PATH, "wb") as onnx_file:
        onnx_file.write(onnx_model_proto.SerializeToString())
    onnx_export_status = "success"
    print(f"Model saved (ONNX):   {FINAL_CLS_ONNX_PATH}")
except ImportError:
    onnx_export_status = "skipped_skl2onnx_not_installed"
    print("ONNX export skipped: skl2onnx not installed. Run: pip install skl2onnx")
except Exception as onnx_error:
    onnx_export_status = f"failed: {str(onnx_error)}"
    print(f"ONNX export failed: {onnx_error}")

final_cls_metadata["onnx_export_status"] = onnx_export_status

with open(FINAL_CLS_METADATA_PATH, "w", encoding="utf-8") as meta_file:
    json.dump(final_cls_metadata, meta_file, ensure_ascii=False, indent=2, default=str)
print(f"Metadata saved:       {FINAL_CLS_METADATA_PATH}")

final_json["final_model_production"] = {
    "best_model_name": best_final_model_name,
    "selection_criterion": "highest average test ROC AUC across all splits",
    "model_avg_roc_auc_per_model": {str(k): float(v) for k, v in model_avg_roc_auc.items()},
    "training_data_size": len(all_labels),
    "training_time_seconds": float(final_cls_train_time),
    "onnx_export_status": onnx_export_status,
    "model_files": {
        "joblib": str(FINAL_CLS_JOBLIB_PATH),
        "onnx": str(FINAL_CLS_ONNX_PATH),
        "metadata_json": str(FINAL_CLS_METADATA_PATH),
    },
}

with open(OUTPUT_JSON_PATH, "w", encoding="utf-8") as output_file:
    json.dump(final_json, output_file, ensure_ascii=False, indent=2, default=str)

print(f"\nUpdated results JSON: {OUTPUT_JSON_PATH}")
print("\n" + "=" * 65)
print("ALL EXPORTS COMPLETE")
print("=" * 65)
print(f"  Classifier (joblib):  {FINAL_CLS_JOBLIB_PATH.name}")
print(f"  Classifier metadata:  {FINAL_CLS_METADATA_PATH.name}")
print(f"  ONNX export status:   {onnx_export_status}")

